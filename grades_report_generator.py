import tkinter as tk
from tkinter import filedialog
import ttkbootstrap as tb
from ttkbootstrap import Style, Button, Radiobutton, Label, Entry, Checkbutton, Combobox, LabelFrame
from ttkbootstrap.dialogs import Messagebox
import threading
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import time
from PIL import Image, ImageTk
import ctypes
from bs4 import BeautifulSoup
import time
from datetime import datetime
import fnmatch
from openpyxl import Workbook
from openpyxl.drawing.image import Image as opimage

class GradesReportGeneratorApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Grades Report Generator")
        self.geometry('300x450')
        self.resizable(False, False)
        self.style = Style(theme="superhero")
        # Add an icon for the task bar and window
        myappid = 'mycompany.myproduct.subproduct.version'  # ramdom string
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

        # Add an icon for the window
        img_path = 'grades_scraper-logo-ico.ico'
        self.iconbitmap(img_path)
        self.iconbitmap(default=img_path)
        self.driver = None
        self.create_login_widgets()
        self.center_window()

        # Create attributes to hold the Select elements
        self.select_status_turma = None
        self.select_semester = None
        self.select_status_aluno = None
        self.select_teacher = None

        # Initialize a flag to tell if the scraping is happening and be able to stop it
        self.stop_flag = False

        # Initialize variables to keep track of time
        self.start_time = None
        self.is_scraping = False

    # Center window and make sure it can't be resized
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x_offset = (self.winfo_screenwidth() - width) // 2
        y_offset = (self.winfo_screenheight() - height) // 2
        self.geometry(f'{width}x{height}+{x_offset}+{y_offset}')

    # Create the login window
    def create_login_widgets(self):
        self.logo_frame = tk.Frame(self)
        self.logo_frame.pack()

        self.logo_image = Image.open("grades_scraper logo (1).png")
        self.logo_image = self.logo_image.resize((200, 90), Image.LANCZOS)  # Resampling.LANCZOS
        self.logo_image = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = Label(self.logo_frame, image=self.logo_image)
        self.logo_label.pack()

        self.login_frame = tk.Frame(self)
        self.login_frame.pack()

        self.login_label = Label(self.login_frame, text="E-mail:", font=("Helvetica", 20))
        self.login_label.grid(row=0, column=0, columnspan=2, padx=10, pady=5)

        self.login_entry = Entry(self.login_frame, width=15)
        self.login_entry.insert(tk.END, "Email")
        self.login_entry.grid(row=1, column=0, padx=1, pady=5, sticky="e")

        self.email_suffix_label = tb.Label(self.login_frame, text="@cna.com.br", font=("Helvetica", 10),
                                           bootstyle="secondary")
        self.email_suffix_label.grid(row=1, column=1, padx=1, pady=5, sticky="w")

        self.login_entry.bind("<Button-1>", lambda a: self.login_entry.delete(0, tk.END))
        self.login_entry.bind("<KeyRelease>", lambda a: self.check_entry())

        self.pass_label = Label(self.login_frame, text="Password:", font=("Helvetica", 20))
        self.pass_label.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

        # Make the password show as hidden characters
        self.pass_entry = Entry(self.login_frame, show='*', width=30)
        self.pass_entry.insert(tk.END, "Password")
        self.pass_entry.grid(row=3, column=0, columnspan=2, padx=10, pady=5)
        # Bind the Enter key to the login button
        self.pass_entry.bind("<Button-1>", lambda a: self.pass_entry.delete(0, tk.END))
        self.pass_entry.bind("<KeyRelease>", lambda a: self.check_entry())
        # Create and cofigure the show password button
        self.remember_var = tk.BooleanVar()
        self.pass_check = tb.Checkbutton(self.login_frame,
                                         text="Show password",
                                         variable=self.remember_var,
                                         onvalue=True,
                                         offvalue=False,
                                         command=self.toggle_password,
                                         bootstyle="primary, round-toggle")
        self.pass_check.grid(row=4, column=0, columnspan=2, padx=10, pady=5)
        # Disable the login button after clicking it
        self.login_button = Button(self.login_frame, text="Log in", padding=5, width=20, state="disabled",
                                   command=self.login)
        self.login_button.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

        self.status_label = tb.Label(self.login_frame, text="", bootstyle="warning")
        self.status_label.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

        self.progressbar = tb.Progressbar(self.login_frame, orient=tk.HORIZONTAL, length=200, mode='determinate', maximum=100, bootstyle="primary")

        self.sign_label = tb.Label(text="designed by Vinny", bootstyle="secondary")
        self.sign_label.pack(anchor="se", padx=5, ipady=5, side="bottom")
        # Close the window
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def toggle_password(self):
        if self.remember_var.get():
            self.pass_entry.config(show='')
            self.pass_check.config(text='Hide password')
        else:
            self.pass_entry.config(show='*')
            self.pass_check.config(text='Show password')

    def check_entry(self):
        if self.login_entry.get() == "Email" or self.pass_entry.get() == "Password":
            self.login_button.config(state='disabled')
        elif self.login_entry.get() == "" or self.pass_entry.get() == "":
            self.login_button.config(state='disabled')
        else:
            self.login_button.config(state='normal')
            # Bind the enter key to click on the login button
            self.bind('<Return>', self.login)

    def login(self, event=None):
        self.login_button.config(state="disabled")
        self.status_label.config(text="Loading...", bootstyle="warning")
        self.progressbar.grid(row=7, column=0, columnspan=2, padx=10, pady=5)
        self.progressbar["value"] = 0
        self.update()

        try:
            self.login_button.config(state="disabled")
            self.status_label.config(text="Opening browser...", bootstyle="warning")
            self.progressbar["value"] += 20
            self.update()
            service = Service()
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')
            options.add_experimental_option('detach', True)
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--disable-gpu')
            self.driver = webdriver.Chrome(options=options, service=service)
            global wait
            global wait2
            wait = WebDriverWait(self.driver, 30)
            wait2 = WebDriverWait(self.driver, 6)

            self.driver.get("https://www.cnabox.com.br/Account/Login")
            self.driver.maximize_window()

            login = wait.until(EC.element_to_be_clickable((By.ID, "Email")))
            login.send_keys(self.login_entry.get() + "@cna.com.br")
            login.submit()
            self.status_label.config(text="Typing e-mail...", bootstyle="warning")
            self.progressbar["value"] += 20
            self.update()

            ps = wait.until(EC.element_to_be_clickable((By.ID, "Password")))
            ps.send_keys(self.pass_entry.get())
            ps.submit()
            self.status_label.config(text="Typing password...", bootstyle="warning")
            self.progressbar["value"] += 20
            self.update()

            self.status_label.config(text="Closing popups", bootstyle="warning")
            self.progressbar["value"] += 20
            self.update()

            try:
                error_message = wait2.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="login-form"]/div/div/div/ul/li')))
                if error_message.is_displayed():
                    self.status_label.config(text="Invalid email or password. Please try again.", bootstyle="danger")
                    self.progressbar.stop()
                    self.driver.quit()
                    return
            except Exception as e:
                print(e)
                pass

            if self.driver.current_url == "https://www.cnabox.com.br/#/Dashboard":
                # Close the bazzilion speech bubbles from the chat that block clicking the buttons later
                # try:
                #     message1_xpath = '//*[@id="omnichat-message-popup"]'
                #     close_message1 = wait2.until(EC.visibility_of_element_located((By.XPATH, message1_xpath)))
                #     close_message1.click()
                #     time.sleep(1)
                # except Exception as e:
                #     print(e)
                #     pass
                #
                # try:
                #     message2_xpath = '//*[@id="bigBoxColor1"]/div[1]/a[2]'
                #     close_message2 = wait2.until(EC.element_to_be_clickable((By.XPATH, message2_xpath)))
                #     close_message2.click()
                # except Exception as e:
                #     print(e)
                #     pass

                popup = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="detalhes-versao-box"]/div/div/div[1]/button')))
                popup.send_keys(Keys.ESCAPE)
                self.progressbar["value"] += 20
                self.status_label.config(text="Login successful!", bootstyle="success")
                self.update()
                time.sleep(2)
                self.clear_login_widgets()
                self.create_widgets_after_login()
            else:
                self.status_label.config(text="Login failed. Please try again.", bootstyle="danger")
                self.progressbar.stop()
                self.driver.quit()
        except ConnectionRefusedError:
            self.status_label.config(text="Connection failed. Please try again.", bootstyle="danger")
            self.progressbar.stop()
            self.driver.quit()
        except Exception as e:
            self.status_label.config(text="An error occurred. Please try again.", bootstyle="danger")
            self.progressbar.stop()
            self.driver.quit()
            print(e)

    def clear_login_widgets(self):
        self.logo_frame.destroy()
        self.login_frame.destroy()
        self.status_label.destroy()
        self.progressbar.destroy()

    def create_widgets_after_login(self):
        self.geometry('600x200')
        self.center_window()
        self.logo_frame = tk.Frame(self)
        self.logo_frame.pack(fill="x", anchor="ne")
        self.logo_image = Image.open("grades_scraper logo (1).png")
        self.logo_image = self.logo_image.resize((100, 45), Image.LANCZOS)
        self.logo_image = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = Label(self.logo_frame, image=self.logo_image)
        self.logo_label.pack(anchor="ne", padx=5, pady=5)

        self.format_frame = LabelFrame(self, text="Select the Format")
        self.format_frame.pack(fill="x", padx=10, pady=20, ipady=5)

        self.format_frame.grid_rowconfigure(0, weight=1)
        self.format_frame.grid_columnconfigure(0, weight=1)
        self.format_frame.grid_columnconfigure(1, weight=1)

        self.tk_format = tk.StringVar()
        self.reg_radio = tb.Radiobutton(self.format_frame, text="Regular",
                                     variable=self.tk_format,
                                     value="regular",
                                     bootstyle="primary toolbutton outline",
                                        command=self.select_format)
        self.reg_radio.grid(column=0, row=0, columnspan=1, padx=20, pady=5, ipadx=50)
        self.conecta_radio = tb.Radiobutton(self.format_frame, text="Conecta",
                                         variable=self.tk_format,
                                         value="conecta",
                                         bootstyle="primary toolbutton outline",
                                            command=self.select_format)
        self.conecta_radio.grid(column=1, row=0, columnspan=1, padx=20, pady=5, ipadx=50)
        self.update()

        # Create the situation frame
        self.situation_frame = LabelFrame(self, text="Select the Situation")

    def select_format(self):
        self.geometry('600x300')
        self.center_window()
        if not hasattr(self, 'situation_frame_created'):
            self.situation_frame_created = False

        if self.situation_frame_created:
            self.clear_situation_frame()

        selection = self.tk_format.get()
        if selection == "conecta":
            self.clear_situation_frame()
            self.driver.get("https://www.cnabox.com.br/#/Alunos")
            box_format = WebDriverWait(self.driver, 30).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="FormatoAulaIds"]')))
            selectFormat = Select(box_format)
            selectFormat.select_by_value("3")
            self.select_situation_conecta()
        elif selection == "regular":
            self.clear_situation_frame()
            self.driver.get("https://www.cnabox.com.br/#/Turma")
            self.select_situation_reg()

    def clear_situation_frame(self):
        for widget in self.situation_frame.winfo_children():
            widget.destroy()

    def clear_format_frame(self):
        for widget in self.format_frame.winfo_children():
            widget.destroy()
        self.format_frame.destroy()

    def update_selection(self, index, var):
        if self.select_status_aluno:
            if var.get():
                self.select_status_aluno.select_by_index(index)
            else:
                self.select_status_aluno.deselect_by_index(index)

    def start_scraping(self):
        self.geometry('1000x550')
        self.center_window()
        self.clear_format_frame()
        self.clear_situation_frame()
        self.situation_frame.pack_forget()
        self.situation_frame.pack(fill="both", padx=10, pady=20, ipady=5)
        self.situation_frame.config(text="Progress")

        self.metersize = 250

        self.group_meter = tb.Meter(self.situation_frame, metersize=self.metersize, subtext="Group", padding=10)
        self.group_meter.grid(row=0, column=0, padx=10)

        self.student_meter = tb.Meter(self.situation_frame, metersize=self.metersize, subtext="Student", padding=10)
        self.student_meter.grid(row=0, column=1, padx=10)

        self.evaluation_meter = tb.Meter(self.situation_frame, metersize=self.metersize, subtext="Evaluation", padding=10)
        self.evaluation_meter.grid(row=0, column=2, padx=10)

        self.elapsed_time_var = tk.StringVar()
        self.elapsed_time_label = tb.Label(self.situation_frame, textvariable=self.elapsed_time_var, anchor="center", bootstyle="light", font=("helvetica", 15))
        self.elapsed_time_label.grid(row=1, column=0, columnspan=3, pady=10)

        self.stop_button = tb.Button(self.situation_frame, bootstyle="danger", width=20, text="Stop Generating", command=self.stop_scraping)
        self.stop_button.grid(row=2, column=1, pady=10, sticky="nesw")

        self.elapsed_time_var.set("Elapsed Time: 00:00:00")
        self.start_time = time.time()
        self.is_scraping = True
        self.update_elapsed_time()
        # Start the scraping process in a separate thread
        threading.Thread(target=self.scrape_data).start()

    def set_meter_subtext(self, meter, text):
        font_size = self.calculate_font_size(text)
        meter.config(subtext=text, subtextfont=("Arial", font_size))

    def calculate_font_size(self, text):
        base_size = 10  # Base font size
        max_length = 10  # Maximum length of text for base size
        if len(text) > max_length:
            return int(base_size * (max_length / len(text)))
        return base_size

    def scrape_data(self):
        # Start the scraping process
        self.update()
        xpath_filter_conecta = '// *[ @ id = "btnAlunosFilter"]'
        xpath_filter_regular = '// *[ @ id = "btnFiltrarTurmas"]'

        if self.driver.current_url == "https://www.cnabox.com.br/#/Turma":
            filter_xpath = xpath_filter_regular
        else:
            filter_xpath = xpath_filter_conecta

        filter = wait.until(EC.element_to_be_clickable((By.XPATH, filter_xpath)))
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'})", filter)
        filter.click()
        # Wait for the page to load
        time.sleep(5)
        # wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="content"]/div[2]/ul/li[2]/span')))
        # Get the html and connect to the BeautifulSoup library to scrape
        pag_turmas = self.driver.page_source
        site = BeautifulSoup(pag_turmas, 'html.parser')
        # Check the html to see where the data is before scraping
        # print(site.prettify())
        # Get all the links from the groups page
        links = [node.get('href') for node in site.find_all("a")]
        # See all the links in the page
        # print(links)
        # Create an empty list to separate and store only the links needed
        partial_links = []
        # Separate and append the links needed to the list
        for link in links:
            if "#/Turma/Details/" in link:
                partial_links.append(link)
        # print(partial_links)
        # Join the partial links in the list to the base url
        base_url = "https://www.cnabox.com.br/"
        full_links = [base_url + x for x in partial_links]
        # See if the links got joined correctly
        print(full_links)
        # Check if there is more than one page of results
        page_number = 3
        while True:
            try:
                # next_button_xpath = f'//*[@id="GridIndexTurmas_paginate"]/ul/li[{page_number}]/a'
                # next_button_box_xpath = f'//*[@id="GridIndexTurmas_paginate"]/ul/li[{page_number}]'
                # next_button_box = wait.until(EC.visibility_of_element_located((By.XPATH, next_button_box_xpath)))
                next_button_xpath = f'//*[@id="GridIndexTurmas_paginate"]/ul/li[{page_number}]/a'
                next_button = self.driver.find_element(By.XPATH, next_button_xpath)
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'})",
                                           next_button)
                if next_button.is_displayed():
                    next_button.click()
                    self.driver.implicitly_wait(10)
                    # Wait for the page to load after clicking
                    pag_turmas2 = self.driver.page_source
                    site2 = BeautifulSoup(pag_turmas2, 'html.parser')
                    links2 = [node.get('href') for node in site2.find_all("a")]
                    partial_links2 = []
                    for link in links2:
                        if "#/Turma/Details/" in link:
                            partial_links2.append(link)
                    base_url = "https://www.cnabox.com.br/"
                    full_links2 = [base_url + x for x in partial_links2]
                    full_links.extend(full_links2)
                    page_number += 1
                else:
                    # If the button is not displayed, break the loop
                    print("No more pages.")
                    break
            except NoSuchElementException:
                # If the "Next" button is not found, exit the loop
                print(f"No more pages or pagination button not found at li[{page_number}].")
                break

        # Count the amount of groups to display to the user
        group_count = len(full_links)
        print(group_count)
        if group_count > 0:
            self.group_meter.configure(amounttotal=group_count, amountused=0, textright="out of " + str(group_count), stripethickness=int(self.metersize/group_count))
            self.update()
        else:
            self.group_meter.configure(amounttotal=group_count, amountused=0)
            self.update()
        # Scraping the grades chunk

        # Creating variables to identify the type of group
        Adult = ["TEENS*", "BAS*", "INT*", "PADV*", "ADV*", "MAST*"]
        A1 = ["A1*"]
        Teen_1 = ["Teen_UP1*"]
        Teen_2 = ["Teen_UP2*"]
        Teen_3_4 = ["Teen_UP3*", "Teen_UP4*"]
        Kids = ["KIDS*"]
        Fun = ["YK*"]
        Garden = ["PS*"]
        Yard_1_2 = ["LK1*", "LK2*"]
        Yard_3_4 = ["LK3*", "LK4*"]
        Espanõl = ["ESP*"]
        Joven = ["*EJ*"]

        # Create the headers for adult
        headers_adult = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "CONTINUOUS PERFORMANCE - MID",
            "CONTINUOUS PERFORMANCE - FINAL",
            "CONSOLIDATION ACTIVITIES - U1",
            "CONSOLIDATION ACTIVITIES - U2",
            "CONSOLIDATION ACTIVITIES - U3",
            "CONSOLIDATION ACTIVITIES - U4",
            "CONSOLIDATION ACTIVITIES - U5",
            "CONSOLIDATION ACTIVITIES - U6",
            "CONSOLIDATION ACTIVITIES - U7",
            "CONSOLIDATION ACTIVITIES - U8",
            "WEB LESSONS - U1",
            "WEB LESSONS - U2",
            "WEB LESSONS - U3",
            "WEB LESSONS - U4",
            "WEB LESSONS - U5",
            "WEB LESSONS - U6",
            "WEB LESSONS - U7",
            "WEB LESSONS - U8",
            "ORAL TEST - MID",
            "ORAL TEST - FINAL",
            "WRITTEN TEST - MID",
            "WRITTEN TEST - FINAL",
        ]

        # Create the headers for A1
        headers_a1 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "PORTFOLIO - UNIT 1 - U1",
            "PORTFOLIO - UNIT 2 - U2",
            "PORTFOLIO - UNIT 3 - U3",
            "PORTFOLIO - UNIT 4 - U4",
            "PORTFOLIO - UNIT 5 - U5",
            "PORTFOLIO - UNIT 6 - U6",
            "DIGITAL CONTENT - UNIT 1 - U1",
            "DIGITAL CONTENT - UNIT 2 - U2",
            "DIGITAL CONTENT - UNIT 3 - U3",
            "DIGITAL CONTENT - UNIT 4 - U4",
            "DIGITAL CONTENT - UNIT 5 - U5",
            "DIGITAL CONTENT - UNIT 6 - U6",
            "WRITTEN TEST - FINAL",
        ]

        # Create the headers for teen
        headers_teen_1 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "WRITTEN PRODUCTION - MIDTERM -",
            "WRITTEN PRODUCTION - FINAL -",
            "ORAL PRODUCTION - MIDTERM -",
            "ORAL PRODUCTION - FINAL -",
            "WRITTEN TEST -",
            "PROJECT 1 - UN 1, 2 E 3",
            "PROJECT 2 - UN 4, 5 E 6",
            "DIGITAL CONTENT - WELCOME LESSON - 0",
            "DIGITAL CONTENT - UNIT 1 - U1",
            "DIGITAL CONTENT - UNIT 2 - U2",
            "DIGITAL CONTENT - UNIT 3 - U3",
            "DIGITAL CONTENT - CONSOLIDATION 1 - C1",
            "DIGITAL CONTENT - UNIT 4 - U4",
            "DIGITAL CONTENT - UNIT 5 - U5",
            "DIGITAL CONTENT - UNIT 6 - U6",
            "DIGITAL CONTENT - CONSOLIDATION 2 - C2",
        ]

        headers_teen_2 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "WRITTEN PRODUCTION - MIDTERM -",
            "WRITTEN PRODUCTION - FINAL -",
            "ORAL PRODUCTION - MIDTERM -",
            "ORAL PRODUCTION - FINAL -",
            "WRITTEN TEST -",
            "PROJECT 1 - UN 1, 2 E 3",
            "PROJECT 2 - UN 4, 5 E 6",
            "DIGITAL CONTENT - WELCOME LESSON -",
            "DIGITAL CONTENT - UNIT 1 - U1",
            "DIGITAL CONTENT - UNIT 2 - U2",
            "DIGITAL CONTENT - UNIT 3 - U3",
            "DIGITAL CONTENT - CONSOLIDATION 1 - C1",
            "DIGITAL CONTENT - UNIT 4 - U4",
            "DIGITAL CONTENT - UNIT 5 - U5",
            "DIGITAL CONTENT - UNIT 6 - U6",
            "DIGITAL CONTENT - CONSOLIDATION 2 - C2",
        ]

        headers_teen_3_4 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "WRITTEN PRODUCTION - MIDTERM -",
            "WRITTEN PRODUCTION - FINAL -",
            "ORAL PRODUCTION - MIDTERM -",
            "ORAL PRODUCTION - FINAL -",
            "WRITTEN TEST - FINAL",
            "PROJECT 1 -",
            "PROJECT 2 -",
            "DIGITAL CONTENT - WELCOME LESSON -",
            "DIGITAL CONTENT - UNIT 1 - U1",
            "DIGITAL CONTENT - UNIT 2 - U2",
            "DIGITAL CONTENT - UNIT 3 - U3",
            "DIGITAL CONTENT - CONSOLIDATION 1 - C1",
            "DIGITAL CONTENT - UNIT 4 - U4",
            "DIGITAL CONTENT - UNIT 5 - U5",
            "DIGITAL CONTENT - UNIT 6 - U6",
            "DIGITAL CONTENT - CONSOLIDATION 2 - C2",
        ]

        # Create the headers for kids
        headers_kids = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "CONTINUOUS PERFORMANCE - MID",
            "CONTINUOUS PERFORMANCE - FINAL",
            "CONSOLIDATION ACTIVITIES - U1",
            "CONSOLIDATION ACTIVITIES - U2",
            "CONSOLIDATION ACTIVITIES - U3",
            "CONSOLIDATION ACTIVITIES - U4",
            "CONSOLIDATION ACTIVITIES - U5",
            "CONSOLIDATION ACTIVITIES - U6",
            "CONSOLIDATION ACTIVITIES - U7",
            "CONSOLIDATION ACTIVITIES - U8",
            "CONSOLIDATION ACTIVITIES - U9",
            "CONSOLIDATION ACTIVITIES - U10",
            "CONSOLIDATION ACTIVITIES - U11",
            "CONSOLIDATION ACTIVITIES - U12",
            "CONSOLIDATION ACTIVITIES - U13",
            "CONSOLIDATION ACTIVITIES - U14",
            "WEB LESSONS - WELCOME",
            "WEB LESSONS - U1",
            "WEB LESSONS - U2",
            "WEB LESSONS - U3",
            "WEB LESSONS - U4",
            "WEB LESSONS - U5",
            "WEB LESSONS - U6",
            "WEB LESSONS - U7",
            "WEB LESSONS - U8",
            "WEB LESSONS - U9",
            "WEB LESSONS - U10",
            "WEB LESSONS - U11",
            "WEB LESSONS - U12",
            "WEB LESSONS - U13",
            "WEB LESSONS - U14",
            "ORAL TEST - MID",
            "ORAL TEST - FINAL",
            "WRITTEN TEST - MID",
            "WRITTEN TEST - FINAL",
        ]

        # Create the headers for fun
        headers_fun = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "ASSESSMENT - 1° BIMESTRE",
            "ASSESSMENT - 2° BIMESTRE",
            "GLOBAL ASSESSMENT - TOTAL",
            "DESEMPENHO ORAL - 1° BIMESTRE",
            "DESEMPENHO ORAL - 2° BIMESTRE",
            "ATIVIDADES - 1° BIMESTRE",
            "ATIVIDADES - 2° BIMESTRE",
            "WEB LESSONS - WELCOME",
            "WEB LESSONS - U1",
            "WEB LESSONS - U2",
            "WEB LESSONS - U3",
            "WEB LESSONS - U4",
            "WEB LESSONS - U5",
            "WEB LESSONS - U6",
            "WEB LESSONS - U7",
            "WEB LESSONS - U8",
        ]

        # Create the headers for garden
        headers_garden = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "MÉDIA - 1° BIMESTRE",
            "MÉDIA - 2° BIMESTRE",
            "FINAL TEST - TOTAL",
            "DESEMPENHO ORAL 1 - 1° BIMESTRE",
            "DESEMPENHO ORAL 2 - 2° BIMESTRE",
            "ATIVIDADES - 1° BIMESTRE",
            "ATIVIDADES - 2° BIMESTRE",
            "WEB LESSONS - U1",
            "WEB LESSONS - U2",
            "WEB LESSONS - U3",
            "WEB LESSONS - U4",
            "WEB LESSONS - U5",
            "WEB LESSONS - U6",
        ]

        # Create the headers for yard12
        headers_yard12 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "GLOBAL ASSESSMENT - TOTAL",
            "DESEMPENHO ORAL 1 - 1° BIMESTRE",
            "DESEMPENHO ORAL 2 - 2° BIMESTRE",
            "ACTIVITIES 1 - 1° BIMESTRE",
            "ACTIVITIES 2 - 2° BIMESTRE",
        ]

        # Create the headers for yard34
        headers_yard34 = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "ASSESSMENT - 1° BIMESTRE",
            "ASSESSMENT - 2° BIMESTRE",
            "GLOBAL ASSESSMENT - TOTAL",
            "DESEMPENHO ORAL 1 - 1° BIMESTRE",
            "DESEMPENHO ORAL 2 - 1° BIMESTRE",
            "DESEMPENHO ORAL 3 - 2° BIMESTRE",
            "DESEMPENHO ORAL 4 - 2° BIMESTRE",
            "ACTIVITIES 1 - 1° BIMESTRE",
            "ACTIVITIES 2 - 1° BIMESTRE",
            "ACTIVITIES 3 - 2° BIMESTRE",
            "ACTIVITIES 4 - 2° BIMESTRE",
        ]

        # Create the headers for spanish
        headers_spanish = [
            "STUDENT",
            "GROUP",
            "APROVEITAMENTO",
            "FREQUÊNCIA",
            "DESEMPEÑO - MITAD",
            "DESEMPEÑO - FINAL",
            "TAREA DE CASA - U1",
            "TAREA DE CASA - U2",
            "TAREA DE CASA - U3",
            "TAREA DE CASA - U4",
            "TAREA DE CASA - U5",
            "TAREA DE CASA - U6",
            "TAREA DE CASA - U7",
            "TAREA DE CASA - U8",
            "ACTIVIDADES EN LA RED - U1",
            "ACTIVIDADES EN LA RED - U2",
            "ACTIVIDADES EN LA RED - U3",
            "ACTIVIDADES EN LA RED - U4",
            "ACTIVIDADES EN LA RED - U5",
            "ACTIVIDADES EN LA RED - U6",
            "ACTIVIDADES EN LA RED - U7",
            "ACTIVIDADES EN LA RED - U8",
            "PRUEBA ORAL",
            "PRUEBA ESCRITA - MITAD",
            "PRUEBA ESCRITA - FINAL",
        ]
        # Create the headers for spanish teen
        headers_spanish_teen = [
            "Desempeño 1er mes - 1 er mes",
            "Desempeño 2er mes - 2er mes",
            "Proyectos",
            "Desempeño 3er mes - 3er mes",
            "Desempeño 4to mes - 4to mes",
            "Proyectos",
            "TAREA DE CASA - U1",
            "TAREA DE CASA - U2",
            "TAREA DE CASA - U3",
            "TAREA DE CASA - U4",
            "TAREA DE CASA - U5",
            "TAREA DE CASA - U6",
            "TAREA DE CASA - U7",
            "TAREA DE CASA - U8",
            "PRUEBA ORAL",
            "PRUEBA ESCRITA - MITAD",
            "PRUEBA ESCRITA - FINAL",
        ]

        # Dictionaries to store scraped data
        scraped_data = {"Adult": {}, "A1": {}, "Teen_1": {}, "Teen_2": {}, "Teen_3_4":{}, "Kids": {}, "Fun": {}, "Garden": {}, "Yard_1_2": {}, "Yard_3_4": {}, "Espanõl": {}, "Joven": {}}

        # delete all but one group to test faster
        # del full_links[:42]

        # Opening each group link
        for group_index, group in enumerate(full_links):
            if self.stop_flag:
                break
            groups = self.driver.get(group)
            group_code = wait.until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="turmaContent"]/div[1]/div[1]/h1/span'))).text.strip("> ")
            print(group_code)
            if group_index > 0:
                self.group_meter.configure(subtext=group_code, amountused=group_index + 1)
                self.update()
            else:
                self.group_meter.configure(subtext=group_code, amountused=group_index)
                self.update()
            if any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Adult):
                group_type = "Adult"
                evaluation_names = headers_adult
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in A1):
                group_type = "A1"
                evaluation_names = headers_a1
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Teen_1):
                group_type = "Teen_1"
                evaluation_names = headers_teen_1
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Teen_2):
                group_type = "Teen_2"
                evaluation_names = headers_teen_2
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Teen_3_4):
                group_type = "Teen_3_4"
                evaluation_names = headers_teen_3_4
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Kids):
                group_type = "Kids"
                evaluation_names = headers_kids
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Fun):
                group_type = "Fun"
                evaluation_names = headers_fun
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Garden):
                group_type = "Garden"
                evaluation_names = headers_garden
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Yard_1_2):
                group_type = "Yard_1_2"
                evaluation_names = headers_yard12
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Yard_3_4):
                group_type = "Yard_3_4"
                evaluation_names = headers_yard34
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Espanõl):
                group_type = "Espanõl"
                evaluation_names = headers_spanish
            elif any(fnmatch.fnmatchcase(group_code, pattern) for pattern in Joven):
                group_type = "Joven"
                evaluation_names = headers_spanish_teen
            else:
                continue

            print("Group Type:", group_type)

            # waiting for the grades table to load and scrolling to it
            aprovacao = wait.until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="turmaContent"]/div[7]/article')))
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'})", aprovacao)
            # time.sleep(1)
            table = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="tabela-turma-aprovacao"]')))
            table2 = table.find_element(By.XPATH, '//*[@id="tabela-turma-aprovacao"]/tbody')
            rows = table2.find_elements(By.TAG_NAME, "tr")
            # Counting the amount of rows (evaluations) to go through them and click
            rows_count = len(rows)
            print(rows_count)
            if rows_count >0:
                self.student_meter.configure(amounttotal=rows_count, amountused=0, textright="out of " + str(rows_count), stripethickness=int(self.metersize/rows_count))
                self.update()
            else:
                self.student_meter.configure(amounttotal=rows_count, amountused=0)
                self.update()

            # Go through each student row and get the names
            x = 1
            for student_index in range(1, rows_count + 1):
                if self.stop_flag:
                    break
                xpath_name = f'//*[@id="tabela-turma-aprovacao"]/tbody/tr[{student_index}]/td[2]'
                name = self.driver.find_element(By.XPATH, xpath_name).text
                print(name)
                self.student_meter.configure(subtext=name, amountused=student_index)
                self.update()

                xpath_aproveitamento = f'//*[@id="tabela-turma-aprovacao"]/tbody/tr[{student_index}]/td[3]'
                aproveitamento = self.driver.find_element(By.XPATH, xpath_aproveitamento).text
                print(aproveitamento)

                xpath_frequencia = f'//*[@id="tabela-turma-aprovacao"]/tbody/tr[{student_index}]/td[4]'
                frequencia = self.driver.find_element(By.XPATH, xpath_frequencia).text
                print(frequencia)

                # Fill the dictionary with student name and group code
                scraped_data[group_type][name] = {"STUDENT": name, "GROUP": group_code, "APROVEITAMENTO": aproveitamento, "FREQUÊNCIA": frequencia}

                # Go through each student, scroll to evaluations window and open it
                xpath_evaluation = f'//*[@id="tabela-turma-aprovacao"]/tbody/tr[{student_index}]/td[3]/a'
                evaluation = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_evaluation)))
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'})", evaluation)
                evaluation.click()
                time.sleep(1)

                try:
                    # Get the grades
                    fieldset_xpath = f'//*[@id="modal-turma-avaliacao-aluno"]/div/div/div[2]/fieldset'
                    fieldset = wait2.until(EC.visibility_of_element_located((By.XPATH, fieldset_xpath)))
                    grades_rows = fieldset.find_elements(By.CLASS_NAME, 'row')
                    grades_rows_count = len(grades_rows)
                    print(grades_rows_count)
                    self.evaluation_meter.configure(amounttotal=grades_rows_count, amountused=0, textright="out of " + str(grades_rows_count), stripethickness=int(self.metersize/grades_rows_count))
                    self.update()

                    # Loop through each evaluation and fill the dictionary
                    for eval_index in range(2, grades_rows_count + 1):
                        if self.stop_flag:
                            break
                        evaluation_xpath = f'//*[@id="modal-turma-avaliacao-aluno"]/div/div/div[2]/fieldset/div[{eval_index}]/section[1]'
                        evaluation_name = self.driver.find_element(By.XPATH, evaluation_xpath).text
                        print(evaluation_name)
                        if group_type == "Fun" or group_type == "Garden" or group_type == "Yard_1_2" or group_type == "Yard_3_4":
                            section = 3
                        else:
                            section = 4
                        self.evaluation_meter.configure(subtext=evaluation_name, amountused=eval_index)
                        self.update()
                        grade_xpath = f'//*[@id="modal-turma-avaliacao-aluno"]/div/div/div[2]/fieldset/div[{eval_index}]/section[{section}]'
                        grade = self.driver.find_element(By.XPATH, grade_xpath).text
                        print(grade)

                        # Convert subjective grades into numerical ones
                        subjective_grades = {
                            "EXCELENTE": "10,00",
                            "BOM": "7,50",
                            "SATISFATÓRIO": "5,00",
                            "MELHORAR": "2,50"
                        }
                        for g in subjective_grades:
                            if g in grade:
                                grade = g
                            else:
                                pass

                        # Fill the dictionary with evaluation name and grade
                        scraped_data[group_type][name][evaluation_name] = grade

                except NoSuchElementException:
                    print("No grades found for", name)

                except TimeoutException:
                    print("No grades found for", name)

                # Close the pop-up window with the grades
                webdriver.ActionChains(self.driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(1)

        # Set the scraping variable to stop the elapsed time
        self.is_scraping = False
        self.stop_button.config(text="Generate other reports", bootstyle="success")
        # self.stop_button.config(command=self.create_widgets_after_login)
        # Quit the WebDriver
        # self.driver.quit()

        # Print the filled dictionary
        print(scraped_data)

        # Create a workbook
        wb = Workbook()

        # Current date and time
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Remove the default sheet
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

        # Map group types to their corresponding headers (skip first two fields: STUDENT and GROUP)
        headers_by_type = {
            "Adult": headers_adult[2:],
            "A1": headers_a1[2:],
            "Teen_1": headers_teen_1[2:],
            "Teen_2": headers_teen_2[2:],
            "Teen_3_4": headers_teen_3_4[2:],
            "Kids": headers_kids[2:],
            "Fun": headers_fun[2:],
            "Garden": headers_garden[2:],
            "Yard_1_2": headers_yard12[2:],
            "Yard_3_4": headers_yard34[2:],
            "Espanõl": headers_spanish[2:],
            "Joven": headers_spanish_teen[2:]
        }

        # Loop over each group type in scraped data
        for group_type, values in scraped_data.items():
            # Create a new worksheet for this group type
            ws = wb.create_sheet(title=group_type)

            # Add logo (image) in the first row
            try:
                logo = opimage('grades_scraper logo (1).png')
                logo.width = 120  # Adjust the width of the image
                logo.height = 50  # Adjust the height of the image
                ws.add_image(logo, 'A1')

                # Get the dimensions of the image
                img_width = logo.width
                img_height = logo.height

                # Adjust the height of the row to fit the image
                ws.row_dimensions[1].height = img_height / 1.2  # Adjust the divisor as needed
            except FileNotFoundError:
                print(f"Logo image not found for sheet '{group_type}', skipping logo...")

            # Write the time and date in the second row
            # ws.merge_cells('A2:F2')
            ws['A2'] = "Report generated at: " + datetime.now().strftime("%H:%M:%S - %d-%m-%Y")

            # Write headers in the third row
            headers = ["STUDENT", "GROUP"] + headers_by_type[group_type]
            ws.append(headers)

            # Write data starting from the fourth row
            for name, data in values.items():
                row = [data["STUDENT"], data["GROUP"]]
                for header in headers[2:]:
                    row.append(data.get(header, ""))
                ws.append(row)

                # Adjust column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width

        # ---- Save Dialog ----
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                title="Grades_report")
        # ---- Save Workbook ----
        if filepath:
            wb.save(filepath)
            print(f"Report successfully generated and saved at: {filepath}")
        else:
            print("Save canceled by the user")

    def stop_scraping(self):
        self.stop_flag = True
        self.is_scraping = False
        self.logo_frame.forget()
        self.clear_situation_frame()
        self.situation_frame.destroy()
        if hasattr(self, "format_frame") and self.format_frame.winfo_exists():
            self.clear_format_frame()
        self.stop_flag = False
        self.create_widgets_after_login()
        self.update()

    def update_elapsed_time(self):
        if self.is_scraping:
            elapsed_time = int(time.time() - self.start_time)
            hours, remainder = divmod(elapsed_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            self.elapsed_time_var.set(f"Elapsed Time: {hours:02}:{minutes:02}:{seconds:02}")
            self.after(1000, self.update_elapsed_time)

    def select_situation_conecta(self):
        self.geometry('600x400')
        self.center_window()
        self.situation_frame.pack(fill="x", padx=10, pady=20, ipady=5)
        self.situation_frame.config(text="Select the Situation")

        select_situation_ele = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="StatusAlunoIds"]')))
        self.select_status_aluno = Select(select_situation_ele)
        options_status_aluno = self.select_status_aluno.options
        values_status_aluno = [option_aluno.get_attribute("value") for option_aluno in options_status_aluno]
        texts_status_aluno = [option_aluno.text for option_aluno in options_status_aluno]

        # Create Checkbuttons for each option
        self.selected_options = []
        for index, text in enumerate(texts_status_aluno):
            aluno_var = tk.IntVar(value=0)
            cb = tb.Checkbutton(self.situation_frame, bootstyle="primary, outline, toolbutton", width=20, text=text, variable=aluno_var, command=lambda idx=index, v=aluno_var: self.update_selection(idx, v))
            cb.grid(row=index // 3, column=index % 3, sticky="nesw", padx=10, pady=5)
            self.selected_options.append((values_status_aluno[index], aluno_var))

            for i in range(3):
                self.situation_frame.grid_columnconfigure(i, weight=1)

            # Call update_selection initially to synchronize with the browser
            if aluno_var.get():
                self.update_selection(index, aluno_var)

        # Add a button to start scraping
        confirm_button = tb.Button(self.situation_frame, text="Generate grades report", bootstyle="success", width=20,
                                   command=self.start_scraping)
        confirm_button.grid(row=(len(texts_status_aluno) // 3) +1, column=1, padx=10, pady=10, sticky="nesw")

    def select_situation_reg(self):
        self.geometry('600x350')
        self.center_window()
        self.situation_frame.pack(fill="x", padx=10, pady=10, ipady=5)
        self.situation_frame.config(text="Select the Situation, Semester and Teacher")

        self.situation_frame.grid_rowconfigure(0, weight=1)
        self.situation_frame.grid_rowconfigure(1, weight=1)
        self.situation_frame.grid_columnconfigure(0, weight=1)
        self.situation_frame.grid_columnconfigure(1, weight=1)
        self.situation_frame.grid_columnconfigure(2, weight=1)

        # First combobox
        select_situation_ele = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="StatusTurmaId"]')))
        select = Select(select_situation_ele)
        options = select.options
        values = [option.get_attribute("value") for option in options]
        texts = [option.text for option in options]

        self.selected_situation_reg = tk.StringVar()
        self.situation_combobox = tb.Combobox(self.situation_frame, textvariable=self.selected_situation_reg)
        self.situation_combobox['values'] = texts
        self.situation_combobox.grid(row=0, column=0, padx=5, pady=5, ipadx=10, sticky="NSEW")

        # Bind the combobox selection to the callback
        self.situation_combobox.bind("<<ComboboxSelected>>",
                                     lambda event: self.on_combobox_select(select, values, texts))

        # Second Combobox
        select_semester_ele = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="TurmaTag"]')))
        select_semester = Select(select_semester_ele)
        semester_options = select_semester.options
        semester_values = [option.get_attribute("value") for option in semester_options]
        semester_texts = [option.text for option in semester_options]

        self.selected_semester_reg = tk.StringVar()
        self.semester_combobox = tb.Combobox(self.situation_frame, textvariable=self.selected_semester_reg)
        self.semester_combobox['values'] = semester_texts
        self.semester_combobox.grid(row=0, column=1, padx=5, pady=5, ipadx=10, sticky="NSEW")

        # Bind the combobox selection to the callback
        self.semester_combobox.bind("<<ComboboxSelected>>",
                                    lambda event: self.on_combobox_select(select_semester, semester_values,
                                                                          semester_texts))

        # third Combobox
        select_teacher_ele = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="ProfessorId"]')))
        select_teacher = Select(select_teacher_ele)
        teacher_options = select_teacher.options
        teacher_values = [option.get_attribute("value") for option in teacher_options]
        teacher_texts = [option.text for option in teacher_options]

        self.selected_teacher = tk.StringVar()
        self.teacher_combobox = tb.Combobox(self.situation_frame, textvariable=self.selected_teacher)
        self.teacher_combobox['values'] = teacher_texts
        self.teacher_combobox.grid(row=0, column=2, padx=5, pady=5, ipadx=10, sticky="NSEW")

        # Bind the combobox selection to the callback
        self.teacher_combobox.bind("<<ComboboxSelected>>",
                                    lambda event: self.on_combobox_select(select_teacher, teacher_values,
                                                                          teacher_texts))

        confirm_button = tb.Button(self.situation_frame, text="Generate grades report", bootstyle="success",
                                   width=30, command=self.start_scraping)
        confirm_button.grid(row=1, column=1, columnspan=1, pady=10, sticky="NSEW")

    def on_combobox_select(self, select, values, texts):
        selected_text = self.selected_situation_reg.get()
        if selected_text in texts:
            selected_index = texts.index(selected_text)
            selected_value = values[selected_index]
            select.select_by_value(selected_value)
        else:
            selected_text = self.selected_semester_reg.get()
            if selected_text in texts:
                selected_index = texts.index(selected_text)
                selected_value = values[selected_index]
                select.select_by_value(selected_value)
            else:
                selected_text = self.selected_teacher.get()
                if selected_text in texts:
                    selected_index = texts.index(selected_text)
                    selected_value = values[selected_index]
                    select.select_by_value(selected_value)

    def get_selected_situation(self):
        selected_text = self.selected_situation_reg.get()
        selected_situation_index = self.situation_combobox['values'].index(selected_text)
        selected_value = self.situation_values[selected_situation_index]
        select_situation_ele = self.wait.until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="StatusTurmaId"]')))
        select_situation = Select(select_situation_ele)
        select_situation.select_by_value(selected_value)

    def on_closing(self):
        # if messagebox.askokcancel("Quit", "Do you want to quit?", icon="warning"):
        if Messagebox.show_question("Do you want to close the app?","Close App") == "Yes":
            if self.driver:
                self.driver.quit()
            self.destroy()


if __name__ == "__main__":
    app = GradesReportGeneratorApp()
    app.mainloop()